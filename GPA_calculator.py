import os
import time
import asyncio
import subprocess
import sys
        

def get_grade(transcript_url,excel_path,username):
    driver = webdriver.Chrome()
    driver.get(transcript_url) 
    isInput = True
    time.sleep(1)
    get_url = driver.current_url
    while (get_url != transcript_url):
        if ("https://sso.hcmut.edu.vn/cas/login" in get_url):
            input_username = driver.find_element(By.ID, 'username')
            input_ = driver.find_element(By.ID, 'password')
            if isInput:
                input_username.send_keys(username)
                input_.send_keys("")
                isInput = False
            username = input_username.get_attribute("value")
            _ = input_.get_attribute("value")   
        time.sleep(1)
        get_url = driver.current_url
        
    time.sleep(7)
    body = driver.find_element(By.TAG_NAME, 'body')
    body.send_keys(Keys.CONTROL, 'a')
    body.send_keys(Keys.CONTROL, 'c')
    driver.quit()

    clipped = pyperclip.paste()
    clipped = clipped.split('\r\n')
    clipped = [item.split('\t') for item in clipped]
    
    wb = Workbook()
    ws = wb.active
    for row, row_data in enumerate(clipped, start=1):
        for col, cell_data in enumerate(row_data, start=1):
            ws.cell(row=row, column=col, value=cell_data)
    wb.save(excel_path)
    
    return username, _


def calculate_grade(transcript_url,excel_path,username):
    if os.path.isfile(excel_path):
        wb = load_workbook(excel_path)
        ws = wb.active
        username = ws['A1'].value
    username, _ = get_grade(transcript_url,excel_path,username)
    
    df = pd.read_excel(excel_path)
    df = df.rename(columns={"Unnamed: 1": "Course", "Unnamed: 2": "Course Name", "Unnamed: 5": "Credit", "Unnamed: 3": "Grade_10", "Unnamed: 4": "Grade"})
    map_grade_4 = {'A+': 4,'A': 4,'B+': 3.5,'B': 3,'C+': 2.5,'C': 2,'D+': 1.5,'D': 1,'F': 0}
    df1 = df.loc[:,["Course", "Course Name", "Credit", "Grade_10", "Grade"]]
    has_grade = df1.loc[df1["Grade"].isin(map_grade_4.keys())]
    has_grade["Grade_4"] = list(map(lambda x: map_grade_4[x], has_grade["Grade"]))
    has_grade = has_grade.sort_values(by=["Course","Grade_4"], ascending=True)
    has_grade = has_grade.drop_duplicates(subset=["Course"],keep="last")
    has_grade = has_grade.reset_index(drop=True)
    has_grade.index += 1
    has_grade["Credit"] = has_grade["Credit"].astype(int)
    has_grade["Grade_10"] = has_grade["Grade_10"].astype(float)
    
    if os.name == 'nt':
        os.system('cls')
    else:
        os.system('clear')
    
    print(has_grade.to_string())

    total_credit = sum(has_grade["Credit"])
    print(f"\nTotal credits: {total_credit}\n")

    total_grade = sum(has_grade["Grade_4"] * has_grade["Credit"])
    average_grade = total_grade / total_credit
    print(f"total_grade: {total_grade}")
    print(f"average_grade: {average_grade}\n")

    total_grade_10 = sum(has_grade["Grade_10"] * has_grade["Credit"])
    average_grade_10 = total_grade_10 / total_credit
    print(f"total_grade_10: {total_grade_10}")
    print(f"average_grade_10: {average_grade_10}\n")
    
    has_grade.to_excel(excel_path)
    
    summary = {}
    summary["Total credits:"] = total_credit
    summary["Total 4.0 scale:"] = total_grade
    summary["CGPA 4.0 scale:"] = average_grade
    summary["Total 10 scale:"] = total_grade_10
    summary["CGPA 10 scale:"] = average_grade_10
    
    position = has_grade.shape[0] + 3
    wb = load_workbook(excel_path)
    ws = wb.active
    for key, value in summary.items():
        ws.cell(row=position, column=1, value=str(key))
        ws.cell(row=position, column=2, value=value)
        position += 1
    ws.cell(row=1, column=1, value=username)
    wb.save(excel_path)
    
    message = f"CGPA 4.0 scale:\n->\t {average_grade}\n\nCGPA 10 scale:\n->\t{average_grade_10}"
    return message, has_grade, summary


def web():
    st.header("GPA Calculator")
    st.markdown("### Hello, this is a website supporting HCMUT students calculating their GPA. I hope it can help you!")
    
    
    
    button_clicked = st.button('Calculate GPA!')
    if button_clicked:
        try: 
            transcript_url = "https://mybk.hcmut.edu.vn/app/sinh-vien/ket-qua-hoc-tap/bang-diem-mon-hoc"
            excel_path = "gc.xlsx"
            df, summary = asyncio.run(calculate(transcript_url,excel_path,""))
            st.subheader("Your Academic Transcript")
            st.dataframe(df.loc[:,["Course", "Course Name", "Credit", "Grade_10", "Grade"]])

            st.markdown("""
                - This is your overall performance:
            """)

            col1, col2, col3 = st.columns(3)
            with col1:
                st.markdown(f"""
                    {list(summary.keys())[0]} {list(summary.values())[0]}       
                """ )
            with col2:
                st.markdown(f"""
                    {list(summary.keys())[1]} {list(summary.values())[1]}\n
                    {list(summary.keys())[3]} {list(summary.values())[3]}     
                """ )
            with col3:
                st.markdown(f"""
                    {list(summary.keys())[2]} {round(list(summary.values())[2],5)}\n
                    {list(summary.keys())[4]} {round(list(summary.values())[4],5)}     
                """ )
        except KeyError:
            pass
            
    st.markdown("Copyright © khoi.nguyenminhcs22@hcmut.edu.vn")


async def calculate(transcript_url,excel_path,username):
    message, df, summary = calculate_grade(transcript_url,excel_path,username)
    await DesktopNotifier().send(
        title="",
        message= message,
        sound=DEFAULT_SOUND,
        urgency=Urgency.Critical,
        buttons=[
            Button(
                title="Mark as read",
                on_pressed=lambda: print("Marked as read"),
            )
        ],
        timeout=5,
    )
    return df, summary

    
if __name__ == "__main__":    
    required = [
    "selenium",
    "openpyxl",
    "desktop-notifier",
    "pyperclip",
    "pandas",
    "streamlit",
    ]
    for package in required:
        try:
            __import__(package.replace("-", "_"))
        except ImportError:
            subprocess.run([sys.executable, "-m", "pip", "install", package])
    from selenium import webdriver
    from selenium.webdriver.common.by import By
    from selenium.webdriver.common.keys import Keys
    from openpyxl import Workbook, load_workbook
    from desktop_notifier import DesktopNotifier, Urgency, Button, DEFAULT_SOUND
    import pyperclip
    import streamlit as st
    import pandas as pd
    print("All required packages installed.")  
    web()
    
    
    